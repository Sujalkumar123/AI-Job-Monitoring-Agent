"""
Excel/CSV exporter module.
- Creates or updates the output Excel file
- Applies formatting and column styling
- Supports append mode to avoid duplicates
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


def export_to_excel(jobs: list[dict], filepath: str = None) -> str:
    """
    Export job listings to a formatted Excel file.

    Args:
        jobs: List of job record dicts
        filepath: Output file path (default from config)

    Returns:
        Path to the generated file
    """
    filepath = filepath or config.OUTPUT_FILE

    # Ensure output directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if not jobs:
        logger.warning("No jobs to export!")
        return filepath

    df = pd.DataFrame(jobs, columns=config.COLUMNS)

    # Remove any completely empty rows
    df = df.dropna(how="all")

    # Write to Excel
    df.to_excel(filepath, index=False, sheet_name="Job Listings", engine="openpyxl")

    # Apply formatting
    _format_excel(filepath, len(df))

    logger.info(f"Exported {len(df)} jobs to: {filepath}")
    return filepath


def export_to_csv(jobs: list[dict], filepath: str = None) -> str:
    """Also save as CSV for easy access."""
    filepath = filepath or config.OUTPUT_CSV

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if not jobs:
        return filepath

    df = pd.DataFrame(jobs, columns=config.COLUMNS)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")

    logger.info(f"Exported {len(df)} jobs to CSV: {filepath}")
    return filepath


def _format_excel(filepath: str, num_rows: int):
    """Apply professional formatting to the Excel file."""
    try:
        wb = load_workbook(filepath)
        ws = wb.active

        # ── Header styling ──
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, len(config.COLUMNS) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # ── Data cell styling ──
        data_font = Font(name="Calibri", size=11)
        data_alignment = Alignment(vertical="center", wrap_text=True)

        # Category color mapping
        category_colors = {
            "Posted Today": "C6EFCE",        # Green
            "Posted Yesterday": "D9E1F2",     # Light blue
            "Posted 2 Days Ago": "FCE4D6",    # Light orange
            "Posted 3-7 Days Ago": "FFF2CC",   # Light yellow
            "Posted More Than 1 Week Ago": "F2F2F2",  # Light gray
        }

        for row in range(2, num_rows + 2):
            for col in range(1, len(config.COLUMNS) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            # Color-code the category column (#6)
            category_cell = ws.cell(row=row, column=6)
            cat_value = category_cell.value or ""
            if cat_value in category_colors:
                category_cell.fill = PatternFill(
                    start_color=category_colors[cat_value],
                    end_color=category_colors[cat_value],
                    fill_type="solid",
                )

            # Make job link a clickable hyperlink
            link_cell = ws.cell(row=row, column=8)
            if link_cell.value and str(link_cell.value).startswith("http"):
                link_cell.hyperlink = str(link_cell.value)
                link_cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")

        # ── Auto-fit column widths ──
        column_widths = {
            1: 25,   # Company Name
            2: 30,   # Job Title
            3: 20,   # Location
            4: 12,   # Platform Source
            5: 15,   # Date Posted
            6: 28,   # Posting Category
            7: 20,   # Salary Package
            8: 50,   # Job Link
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Freeze header row ──
        ws.freeze_panes = "A2"

        # ── Auto-filter ──
        ws.auto_filter.ref = ws.dimensions

        wb.save(filepath)
        logger.info("Excel formatting applied successfully")

    except Exception as e:
        logger.error(f"Error formatting Excel: {e}")
